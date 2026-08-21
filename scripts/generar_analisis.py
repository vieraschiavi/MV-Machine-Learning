"""Arma el Excel de análisis del negocio: calificación, rentabilidad y competencia.

El archivo que sale (`docs/MV-AutoML-Studio-Analisis.xlsx`) es un modelo, no una
foto: la hoja «Supuestos» tiene todos los números editables y el resto de las
hojas los referencian con fórmulas. Cambiar el precio, la conversión o el gasto
en publicidad recalcula los 24 meses de los dos escenarios.

Este guión existe para que el Excel sea reproducible; no hace falta correrlo
para usarlo — el archivo ya está en `docs/`.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SALIDA = Path(__file__).resolve().parents[1] / "docs" / "MV-AutoML-Studio-Analisis.xlsx"

AZUL = "1E3A8A"
AZUL_CLARO = "E8EDFA"
GRIS = "F3F4F6"
VERDE = "DCFCE7"
ROJO = "FEE2E2"
AMBAR = "FEF3C7"

TITULO = Font(bold=True, color="FFFFFF", size=11)
FONDO_TITULO = PatternFill("solid", fgColor=AZUL)
NEGRITA = Font(bold=True)
CHICA = Font(size=9, color="6B7280")
BORDE = Border(*[Side(style="thin", color="D1D5DB")] * 4)
ARRIBA = Alignment(vertical="top", wrap_text=True)


def encabezado(hoja, fila, valores, anchos=None):
    for i, v in enumerate(valores, start=1):
        c = hoja.cell(row=fila, column=i, value=v)
        c.font = TITULO
        c.fill = FONDO_TITULO
        c.alignment = Alignment(vertical="center", wrap_text=True)
    hoja.row_dimensions[fila].height = 30
    if anchos:
        for i, a in enumerate(anchos, start=1):
            hoja.column_dimensions[get_column_letter(i)].width = a


def titulo_hoja(hoja, texto, subtexto=""):
    hoja["A1"] = texto
    hoja["A1"].font = Font(bold=True, size=15, color=AZUL)
    if subtexto:
        hoja["A2"] = subtexto
        hoja["A2"].font = CHICA
        hoja["A2"].alignment = ARRIBA


# ─────────────────────────────────────────────────────────────────────────────
# 1 · Calificación
# ─────────────────────────────────────────────────────────────────────────────
CALIFICACIONES = [
    ("Funcionalidad (motor de ML y ETL)", 8.5,
     "DuckDB fuera de memoria sobre Parquet, tres ventanas de validación con holdout ciego, "
     "auditoría de fuga, calibración isotónica y smearing de Duan, SHAP con dirección por "
     "categoría, importancia por permutación, texto con TF-IDF, tableros con KPI automáticos "
     "y comparación contra mes anterior / año anterior / acumulado.",
     "Series de tiempo con estacionalidad, multiclase con más de 10 clases y un modo de "
     "reentrenamiento programado. Nada de eso está hoy."),
    ("Pruebas y calidad", 8.5,
     "La suite corre en cada integración e incluye lo que casi nadie prueba: el .exe "
     "compilado se ejecuta de verdad en Windows (sube un archivo, entrena, calcula SHAP), "
     "y la firma del servidor web se cruza contra el verificador del programa. Cuatro "
     "errores reales los cazó el CI antes que un cliente.",
     "Falta prueba de interfaz de punta a punta (Playwright) y una prueba de carga con un "
     "dataset de varios GB."),
    ("Seguridad", 7.5,
     "Licencias firmadas con Ed25519 y clave privada fuera del repositorio; token de API "
     "obligatorio en el backend local; conectores SQL de sólo lectura con barrera endurecida "
     "(sin comentarios ni literales antes de analizar, una sola sentencia, lista negra); "
     "precios del lado del servidor; el webhook de pago no le cree al aviso y consulta el "
     "estado real.",
     "El instalador no está firmado digitalmente: Windows SmartScreen va a mostrar una "
     "advertencia hasta juntar reputación (un certificado EV cuesta entre US$ 250 y 400 al "
     "año). Falta límite de intentos en las funciones del sitio, y el panel se protege con "
     "una sola contraseña sin segundo factor. El repositorio todavía es público."),
    ("Distribución e instalación", 7.0,
     "Instalador NSIS asistido: el usuario elige el disco y la carpeta, no se clava en C:. "
     "Crea acceso directo en Escritorio y Menú de inicio, deja desinstalador registrado, e "
     "instala por usuario sin pedir permisos de administrador. Se compila solo en cada "
     "integración y publica dos versiones: demo pública y owner en borrador privado.",
     "Sin firma digital, sin actualización automática dentro del programa y sin versión "
     "para macOS ni Linux. El instalador pesa 375 MB y ocupa cerca de 1,6 GB una vez "
     "instalado, por las bibliotecas de ML."),
    ("Diseño del producto", 7.5,
     "Interfaz coherente en tres idiomas, tableros con filtros, tablas y gráficos "
     "interactivos exportables a Excel y CSV, KPI detectados solos, mapa de correlaciones y "
     "SHAP legibles, espacio de preguntas en lenguaje natural sobre el dataset.",
     "Falta una primera pantalla que guíe al que abre el programa por primera vez, plantillas "
     "por rubro (cobranzas, ventas, rotación) y un modo presentación para mostrarle el "
     "resultado a un gerente."),
    ("Web de venta", 7.0,
     "Sitio en tres idiomas con video narrado por idioma, precios claros, cobro con "
     "MercadoPago del lado del servidor, página de gracias y descarga directa del instalador. "
     "Todo estático: carga rápido y no cuesta nada mantenerlo.",
     "No hay dominio propio, ni prueba social (casos, logos, testimonios), ni contenido para "
     "buscadores, ni medición de visitas. Sin eso el sitio convierte a quien ya venía "
     "decidido, no capta."),
    ("Rentabilidad demostrada", 5.0,
     "El circuito de cobro está armado y probado de punta a punta: se paga, se emite la "
     "licencia, el programa la acepta. El panel muestra clientes, ingresos netos y descargas "
     "en vivo. Los costos fijos son casi cero.",
     "Cero clientes, cero facturación y ningún canal de captación andando. El producto puede "
     "cobrar; todavía no cobró. Esta nota sube sola con las primeras diez ventas."),
]

GENERAL = 7.4


def hoja_calificacion(wb):
    h = wb.create_sheet("Calificación")
    titulo_hoja(h, "Calificación del producto, del 1 al 10",
                "Nota puesta contra lo que hoy está funcionando y probado, no contra lo que "
                "está planeado. La columna «qué falta» es lo que hay que hacer para subir "
                "cada nota, en orden de impacto.")
    encabezado(h, 4, ["Etapa", "Nota", "Por qué esa nota", "Qué falta para subirla"],
               anchos=[34, 8, 68, 62])
    fila = 5
    for etapa, nota, por_que, falta in CALIFICACIONES:
        h.cell(row=fila, column=1, value=etapa).font = NEGRITA
        c = h.cell(row=fila, column=2, value=nota)
        c.font = Font(bold=True, size=12)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = PatternFill("solid", fgColor=VERDE if nota >= 8 else AMBAR if nota >= 7 else ROJO)
        h.cell(row=fila, column=3, value=por_que).alignment = ARRIBA
        h.cell(row=fila, column=4, value=falta).alignment = ARRIBA
        h.row_dimensions[fila].height = 78
        for col in range(1, 5):
            h.cell(row=fila, column=col).border = BORDE
        fila += 1

    fila += 1
    h.cell(row=fila, column=1, value="GENERAL").font = Font(bold=True, size=12)
    c = h.cell(row=fila, column=2, value=GENERAL)
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=AZUL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    h.cell(row=fila, column=3, value=(
        "Producto técnicamente sólido y ya vendible, con la parte comercial sin estrenar. "
        "El motor de ML y las pruebas están por encima del promedio de lo que se vende a este "
        "precio; lo que baja el promedio es todo lo que rodea a la venta: dominio, firma del "
        "instalador, prueba social y captación."
    )).alignment = ARRIBA
    h.cell(row=fila, column=4, value=(
        "Tres cosas, en este orden: (1) las primeras diez ventas, aunque sean a conocidos; "
        "(2) dominio propio y certificado de firma; (3) tres casos reales publicados con "
        "números. Con eso la nota general va a 8,5."
    )).alignment = ARRIBA
    h.row_dimensions[fila].height = 78
    h.freeze_panes = "A5"
    return h


# ─────────────────────────────────────────────────────────────────────────────
# 2 · Supuestos (todo lo editable vive acá)
# ─────────────────────────────────────────────────────────────────────────────
SUPUESTOS = [
    ("PRECIOS Y MEZCLA", None, None, None),
    ("Precio Profesional mensual", 39, "US$", "El del sitio hoy"),
    ("Precio Empresa mensual", 129, "US$", "El del sitio hoy"),
    ("Proporción de clientes en Empresa", 0.20, "", "1 de cada 5 compra el plan caro"),
    ("Ingreso medio por cliente y mes", "=B6*(1-B8)+B7*B8", "US$", "Se calcula solo"),
    ("", None, None, None),
    ("COBRO E IMPUESTOS", None, None, None),
    ("Comisión de MercadoPago", 0.0629, "", "Checkout Pro con acreditación inmediata. El "
     "número real lo vas a ver en el panel, en la diferencia entre bruto y neto"),
    ("Tasa efectiva de impuestos", 0.12, "", "Ver la hoja «Impuestos Uruguay»: 12% es el "
     "escenario intermedio, exportando servicios y con exoneración parcial de IRAE"),
    ("", None, None, None),
    ("CRECIMIENTO Y PERMANENCIA", None, None, None),
    ("Clientes nuevos por mes sin publicidad", 3, "clientes", "Boca a boca, red propia y "
     "quien llega al sitio por su cuenta"),
    ("Crecimiento mensual de ese número", 0.08, "", "8% más cada mes: llegan más porque hay "
     "más clientes hablando del producto"),
    ("Bajas mensuales", 0.05, "", "5% de los activos se va cada mes. En software para "
     "empresas chicas es una cifra normal"),
    ("", None, None, None),
    ("PUBLICIDAD", None, None, None),
    ("Gasto mensual en publicidad", 400, "US$", "Escenario con inversión"),
    ("Costo de conseguir un cliente", 90, "US$", "Lo que cuesta cada venta pagada. Con "
     "producto sin marca conocida, en LATAM, es un supuesto prudente"),
    ("", None, None, None),
    ("COSTOS FIJOS", None, None, None),
    ("Vercel Pro", 20, "US$/mes", "Obligatorio en cuanto el sitio cobre: el plan gratuito "
     "es sólo para proyectos sin fines comerciales"),
    ("Dominio", 2.5, "US$/mes", "US$ 30 al año, prorrateado"),
    ("Correo (Resend)", 0, "US$/mes", "Gratis hasta 3.000 correos por mes"),
    ("Firma digital del instalador", 27, "US$/mes", "Certificado de firma de código: unos "
     "US$ 320 al año. Sin esto Windows muestra una advertencia al instalar"),
    ("Costo fijo total", "=SUM(B25:B28)", "US$/mes", "Se calcula solo"),
]


def hoja_supuestos(wb):
    h = wb.create_sheet("Supuestos")
    titulo_hoja(h, "Supuestos del modelo",
                "Todo lo de esta hoja es editable. Las hojas de rentabilidad lo referencian "
                "con fórmulas: si cambiás un número acá, los 24 meses se recalculan solos.")
    encabezado(h, 4, ["Concepto", "Valor", "Unidad", "De dónde sale"],
               anchos=[38, 14, 12, 66])
    fila = 5
    for concepto, valor, unidad, nota in SUPUESTOS:
        if valor is None and unidad is None:
            if concepto:
                c = h.cell(row=fila, column=1, value=concepto)
                c.font = Font(bold=True, color=AZUL, size=10)
                for col in range(1, 5):
                    h.cell(row=fila, column=col).fill = PatternFill("solid", fgColor=AZUL_CLARO)
            fila += 1
            continue
        h.cell(row=fila, column=1, value=concepto)
        c = h.cell(row=fila, column=2, value=valor)
        c.font = NEGRITA
        if isinstance(valor, float) and 0 < valor < 1:
            c.number_format = "0.0%"
        elif isinstance(valor, (int, float)):
            c.number_format = "#,##0.00"
        h.cell(row=fila, column=3, value=unidad)
        h.cell(row=fila, column=4, value=nota).alignment = ARRIBA
        h.row_dimensions[fila].height = 30
        fila += 1
    h.freeze_panes = "A5"
    return h


# Filas de «Supuestos» a las que apuntan las fórmulas (el orden de SUPUESTOS
# manda; si se agrega una fila arriba, estos números cambian).
S = {
    "arpu": "Supuestos!$B$9",
    "comision": "Supuestos!$B$12",
    "impuestos": "Supuestos!$B$13",
    "nuevos": "Supuestos!$B$16",
    "crecimiento": "Supuestos!$B$17",
    "bajas": "Supuestos!$B$18",
    "ads": "Supuestos!$B$21",
    "cac": "Supuestos!$B$22",
    "fijos": "Supuestos!$B$29",
}

HITOS = {1, 3, 6, 9, 12, 18, 24}


def hoja_escenario(wb, nombre, con_ads: bool):
    h = wb.create_sheet(nombre)
    titulo_hoja(
        h,
        f"Rentabilidad neta mes a mes — {'con' if con_ads else 'sin'} inversión en publicidad",
        "Todas las cifras en dólares. «Neto» es lo que queda después de la comisión de "
        "MercadoPago, los costos fijos, la publicidad y los impuestos. Las filas de los meses "
        "1, 3, 6, 9, 12, 18 y 24 están resaltadas.")

    columnas = ["Mes", "Clientes nuevos", "Bajas", "Clientes activos", "Facturación bruta",
                "Comisión MercadoPago", "Costos fijos"]
    if con_ads:
        columnas.append("Publicidad")
    columnas += ["Resultado antes de impuestos", "Impuestos", "RENTABILIDAD NETA",
                 "Acumulado"]
    encabezado(h, 4, columnas, anchos=[7, 15, 9, 14, 15, 17, 12] +
               ([12] if con_ads else []) + [21, 12, 18, 14])

    primera = 5
    for i in range(24):
        f = primera + i
        mes = i + 1
        ant = f - 1
        h.cell(row=f, column=1, value=mes)

        # Clientes nuevos del mes. Con publicidad, la parte pagada es fija
        # (gasto / costo por cliente) y sólo crece la parte que llega sola.
        if con_ads:
            pagos = f"{S['ads']}/{S['cac']}"
            valor_nuevos = (f"={S['nuevos']}+{pagos}" if i == 0
                            else f"=(B{ant}-{pagos})*(1+{S['crecimiento']})+{pagos}")
        else:
            valor_nuevos = (f"={S['nuevos']}" if i == 0
                            else f"=B{ant}*(1+{S['crecimiento']})")
        h.cell(row=f, column=2, value=valor_nuevos).number_format = "#,##0.0"

        # bajas sobre los activos del mes anterior
        h.cell(row=f, column=3,
               value=(0 if i == 0 else f"=D{ant}*{S['bajas']}")).number_format = "#,##0.0"
        # activos
        h.cell(row=f, column=4,
               value=(f"=B{f}" if i == 0 else f"=D{ant}+B{f}-C{f}")).number_format = "#,##0.0"
        # facturación bruta
        h.cell(row=f, column=5, value=f"=D{f}*{S['arpu']}").number_format = "#,##0"
        # comisión
        h.cell(row=f, column=6, value=f"=-E{f}*{S['comision']}").number_format = "#,##0"
        # fijos
        h.cell(row=f, column=7, value=f"=-{S['fijos']}").number_format = "#,##0"

        col = 8
        if con_ads:
            h.cell(row=f, column=col, value=f"=-{S['ads']}").number_format = "#,##0"
            col += 1

        antes = get_column_letter(col)
        suma = f"=E{f}+F{f}+G{f}" + (f"+H{f}" if con_ads else "")
        h.cell(row=f, column=col, value=suma).number_format = "#,##0"
        col += 1

        imp = get_column_letter(col)
        h.cell(row=f, column=col,
               value=f"=-MAX(0,{antes}{f})*{S['impuestos']}").number_format = "#,##0"
        col += 1

        neta = get_column_letter(col)
        c = h.cell(row=f, column=col, value=f"={antes}{f}+{imp}{f}")
        c.number_format = "#,##0"
        c.font = NEGRITA
        col += 1

        acum = get_column_letter(col)
        h.cell(row=f, column=col,
               value=(f"={neta}{f}" if i == 0
                      else f"={acum}{ant}+{neta}{f}")).number_format = "#,##0"

        if mes in HITOS:
            for j in range(1, col + 1):
                h.cell(row=f, column=j).fill = PatternFill("solid", fgColor=AZUL_CLARO)

    h.freeze_panes = "B5"
    return h


def hoja_resumen(wb):
    h = wb.create_sheet("Resumen 1-24 meses", 0)
    titulo_hoja(h, "Rentabilidad neta en los meses que pediste",
                "Comparación de los dos escenarios en los meses 1, 3, 6, 9, 12, 18 y 24. "
                "«Del mes» es lo que deja ese mes solo; «acumulado» es todo lo ganado desde "
                "el arranque. Los números salen de las hojas «Sin publicidad» y «Con "
                "publicidad» — cambiá los supuestos y esto se mueve.")

    encabezado(h, 4, ["Mes", "Clientes activos (sin ads)", "Neto del mes (sin ads)",
                      "Acumulado (sin ads)", "Clientes activos (con ads)",
                      "Neto del mes (con ads)", "Acumulado (con ads)",
                      "Diferencia acumulada"],
               anchos=[7, 20, 20, 18, 20, 20, 18, 18])

    fila = 5
    for mes in sorted(HITOS):
        origen = 4 + mes
        h.cell(row=fila, column=1, value=mes).font = NEGRITA
        h.cell(row=fila, column=2, value=f"='Sin publicidad'!D{origen}").number_format = "#,##0.0"
        h.cell(row=fila, column=3, value=f"='Sin publicidad'!J{origen}").number_format = "#,##0"
        h.cell(row=fila, column=4, value=f"='Sin publicidad'!K{origen}").number_format = "#,##0"
        h.cell(row=fila, column=5, value=f"='Con publicidad'!D{origen}").number_format = "#,##0.0"
        h.cell(row=fila, column=6, value=f"='Con publicidad'!K{origen}").number_format = "#,##0"
        h.cell(row=fila, column=7, value=f"='Con publicidad'!L{origen}").number_format = "#,##0"
        c = h.cell(row=fila, column=8, value=f"=G{fila}-D{fila}")
        c.number_format = "#,##0"
        c.font = NEGRITA
        for col in range(1, 9):
            h.cell(row=fila, column=col).border = BORDE
        fila += 1

    fila += 2
    h.cell(row=fila, column=1, value="Cómo leerlo").font = Font(bold=True, color=AZUL, size=12)
    fila += 1
    for linea in [
        "· La publicidad da vuelta rápido con estos supuestos: el mes 1 cierra en pérdida "
        "(unos US$ 50 abajo), pero al mes 3 el acumulado ya es positivo y al mes 24 la "
        "diferencia contra no invertir ronda los US$ 35.000. Eso pasa porque un cliente que "
        "cuesta US$ 90 y deja US$ 57 por mes se paga solo en dos meses.",
        "· Ese resultado depende por completo de un número que todavía NO medimos: los US$ 90 "
        "por cliente. Es el supuesto más frágil de todo el modelo. Si en la práctica te sale "
        "US$ 250 —perfectamente posible con un producto sin marca conocida— la publicidad "
        "pasa a perder plata y el escenario sin ads gana. Antes de subir el gasto, gastá "
        "US$ 300 y medí cuántas ventas salieron: recién ahí este número deja de ser una "
        "apuesta.",
        "· La permanencia pesa más que el precio. Bajar las bajas del 5% al 3% mensual sube el "
        "acumulado de 24 meses más que subir el precio un 20%, y no espanta a nadie.",
        "· El plan Empresa es el que mueve la aguja: al costar más del triple, cada punto de "
        "mezcla que se corra hacia Empresa vale mucho más que un cliente nuevo de Profesional.",
        "· Esto es un modelo, no una promesa. El supuesto más frágil es «3 clientes nuevos por "
        "mes sin publicidad»: hasta que no haya un mes real con ventas, es una apuesta.",
    ]:
        c = h.cell(row=fila, column=1, value=linea)
        c.alignment = ARRIBA
        h.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=8)
        h.row_dimensions[fila].height = 30
        fila += 1
    return h


# ─────────────────────────────────────────────────────────────────────────────
# 5 · Impuestos en Uruguay
# ─────────────────────────────────────────────────────────────────────────────
IMPUESTOS = [
    ("Unipersonal — Literal E",
     "Facturación anual chica (hasta el tope que fija la DGI, del orden de UYU 1.100.000 "
     "por año). Se paga un importe fijo mensual por IVA mínimo e IRAE ficto, más los aportes "
     "a BPS del titular.",
     "Muy baja: el fijo mensual ronda el equivalente a US$ 130-180 entre DGI y BPS.",
     "≈ 6-10% de la facturación mientras seas chico; el porcentaje baja cuanto más facturás.",
     "Es donde conviene arrancar. Cuando pases el tope hay que cambiar de régimen sí o sí."),
    ("Sociedad (SAS o SRL) — régimen general",
     "IRAE 25% sobre la renta neta (ingresos menos gastos deducibles), IVA 22% sobre las "
     "ventas dentro de Uruguay, y 7% de IRPF/IRNR cuando se reparten dividendos.",
     "IRAE 25% + IVA 22% en ventas locales + 7% al retirar.",
     "≈ 25-30% sobre la ganancia si vendés dentro del país.",
     "Es el escenario caro. Sólo tiene sentido si la mayor parte de las ventas son locales "
     "y ya son grandes."),
    ("Sociedad exportando servicios",
     "La venta de software a un cliente del exterior, usado en el exterior, se trata como "
     "exportación de servicios: no lleva IVA. Queda el IRAE sobre la renta neta.",
     "Sin IVA en esas ventas. IRAE 25% sobre la ganancia.",
     "≈ 12-18% sobre la facturación, según cuánto gasto tengas para deducir.",
     "Es el escenario más probable para este producto: el mercado natural es LATAM y el "
     "mundo, no sólo Uruguay."),
    ("Sociedad con exoneración de software",
     "Uruguay exonera de IRAE la renta por producción de soporte lógico desarrollado en el "
     "país, cumpliendo requisitos de nexo (proporción del gasto en desarrollo local y "
     "personal calificado) y registro.",
     "IRAE puede quedar cerca de 0 sobre la parte exonerada. IVA exportación sigue sin "
     "aplicar.",
     "≈ 3-8% sobre la facturación, casi todo BPS y gastos administrativos.",
     "Es el mejor escenario y es alcanzable, pero exige cumplir los requisitos formales y "
     "tener el registro al día. Sin contador no lo armes."),
]


def hoja_impuestos(wb):
    h = wb.create_sheet("Impuestos Uruguay")
    titulo_hoja(h, "Cuánto se lleva el fisco, según cómo factures",
                "Cuatro caminos posibles, del más chico al más armado. La columna de la "
                "derecha es la que hay que llevar a la hoja «Supuestos», en «Tasa efectiva "
                "de impuestos».")
    encabezado(h, 4, ["Régimen", "Qué es", "Qué se paga",
                      "Tasa efectiva sobre facturación", "Cuándo conviene"],
               anchos=[30, 58, 42, 26, 52])
    fila = 5
    for regimen, que_es, que_paga, tasa, cuando in IMPUESTOS:
        h.cell(row=fila, column=1, value=regimen).font = NEGRITA
        h.cell(row=fila, column=2, value=que_es).alignment = ARRIBA
        h.cell(row=fila, column=3, value=que_paga).alignment = ARRIBA
        c = h.cell(row=fila, column=4, value=tasa)
        c.font = NEGRITA
        c.alignment = ARRIBA
        h.cell(row=fila, column=5, value=cuando).alignment = ARRIBA
        h.row_dimensions[fila].height = 74
        for col in range(1, 6):
            h.cell(row=fila, column=col).border = BORDE
        fila += 1

    fila += 1
    c = h.cell(row=fila, column=1, value=(
        "ADVERTENCIA. Los porcentajes de esta hoja son órdenes de magnitud para armar el "
        "modelo, no asesoramiento fiscal. Los topes, las tasas y los requisitos de la "
        "exoneración de software cambian, y el que decide qué régimen te corresponde es un "
        "contador con tus números adelante. Antes de facturar la primera venta, sentate una "
        "hora con uno: cuesta poco y te ahorra elegir mal el régimen, que después es caro de "
        "revertir."))
    c.alignment = ARRIBA
    c.fill = PatternFill("solid", fgColor=AMBAR)
    h.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=5)
    h.row_dimensions[fila].height = 62
    return h


# ─────────────────────────────────────────────────────────────────────────────
# 6 · Competencia
# ─────────────────────────────────────────────────────────────────────────────
COMPETENCIA = [
    # (mercado, nombre, qué hace, precio aproximado, fortaleza, debilidad, cómo compite MV)
    ("Uruguay", "Consultoras de datos locales",
     "Proyectos a medida: un analista arma el modelo y entrega un informe.",
     "US$ 3.000 a 15.000 por proyecto",
     "Confianza, trato cara a cara, entienden el negocio local.",
     "No dejan herramienta: cuando termina el proyecto, la empresa vuelve a depender de "
     "ellos para el próximo.",
     "MV cuesta una fracción y queda instalado. Se puede vender incluso a la consultora, "
     "como herramienta para entregar más rápido."),
    ("Uruguay", "Power BI / Tableau con un analista",
     "Tableros y reportes. Predicción, poco y nada.",
     "US$ 14 por usuario/mes + sueldo del analista",
     "Ya está instalado en casi todas las empresas medianas; nadie discute la compra.",
     "Describe el pasado. Para predecir hay que salir de la herramienta.",
     "MV no compite: se enchufa al lado. El discurso es «esto te dice qué va a pasar, no "
     "qué pasó»."),
    ("Uruguay", "Desarrollador freelance",
     "Scripts de Python a medida, entregados como notebook.",
     "US$ 25 a 60 la hora",
     "Barato y flexible.",
     "Sin validación seria, sin control de fuga, sin mantenimiento. El modelo se pudre en "
     "tres meses y nadie lo nota.",
     "Ahí está la ventaja más defendible de MV: tres ventanas de validación y auditoría de "
     "fuga, o sea números en los que se puede confiar."),
    ("LATAM", "Pecan.ai",
     "AutoML predictivo apuntado a marketing y retención, en la nube.",
     "Empresarial, del orden de US$ 2.000+ por mes",
     "Muy pulido, buena narrativa comercial.",
     "Caro para el bolsillo de la empresa mediana de la región, y los datos van a la nube.",
     "MV corre en la máquina del cliente y cuesta 50 veces menos. Para un banco o una "
     "financiera, que los datos no salgan es un argumento de venta enorme."),
    ("LATAM", "Nubes regionales (Vertex AI, SageMaker, Azure ML)",
     "AutoML por consumo dentro de la nube del proveedor.",
     "Por uso: de US$ 50 a miles por mes",
     "Escala infinita, integración con el resto de la nube.",
     "Requiere un equipo que sepa usarlas. La factura es impredecible. Los datos se suben.",
     "MV es precio fijo, sin factura sorpresa, sin subir nada y sin equipo de datos."),
    ("LATAM", "BI regional + consultoría (SAS, Qlik y socios locales)",
     "Licencia más implementación por integrador.",
     "US$ 10.000 a 100.000 por año",
     "Relación instalada con las empresas grandes.",
     "Ciclo de venta larguísimo, implementación de meses.",
     "MV entra por abajo: se instala en una tarde y da resultado el mismo día."),
    ("Mundo", "DataRobot",
     "La plataforma de AutoML empresarial de referencia.",
     "Del orden de US$ 50.000+ por año",
     "Completísima, gobernanza, MLOps, soporte.",
     "Precio de otro planeta para una PyME. Implementación pesada.",
     "No se compite de frente. MV es «el DataRobot que sí podés pagar», con el 20% de las "
     "funciones que el 90% de las empresas usa."),
    ("Mundo", "H2O Driverless AI",
     "AutoML potente, en la nube o instalado.",
     "Decenas de miles de dólares por año",
     "Motor técnicamente muy bueno.",
     "Curva de aprendizaje y precio empresarial.",
     "MV apunta al usuario que no es científico de datos: se elige el objetivo y listo."),
    ("Mundo", "Dataiku",
     "Plataforma colaborativa de datos y ML.",
     "Edición gratuita limitada; empresarial de cinco cifras",
     "Cubre todo el ciclo, muy adoptada.",
     "Necesita un equipo. Pensada para organizaciones con área de datos.",
     "MV es de una sola persona: el gerente que quiere una respuesta, sin equipo."),
    ("Mundo", "KNIME / Orange (código abierto)",
     "Flujos visuales de análisis y ML, gratis.",
     "Gratis el escritorio; Hub pago",
     "Cuestan cero y son muy capaces.",
     "Hay que saber armar el flujo. No traen validación honesta de fábrica ni tableros de "
     "negocio.",
     "El competidor más incómodo, porque el precio es imbatible. MV gana en tiempo hasta el "
     "resultado y en que la validación viene puesta, no hay que saber armarla."),
    ("Mundo", "Obviously.AI, Akkio y similares",
     "AutoML web sin código, por suscripción mensual.",
     "US$ 75 a 1.500 por mes",
     "Simplísimos de usar, muy buen precio de entrada.",
     "En la nube, con topes de filas, y poco control sobre la validación.",
     "Es el competidor más parecido en precio. MV se diferencia en tres cosas: corre local, "
     "no tiene tope de filas en el plan pago, y muestra el holdout ciego."),
    ("Mundo", "Alteryx Designer",
     "Preparación de datos y análisis para el analista de negocio.",
     "≈ US$ 5.000 por usuario y año",
     "Estándar en muchas áreas de finanzas y operaciones.",
     "Caro por usuario; el ML es secundario.",
     "MV cubre la parte de ML a una fracción, y el ETL automático le compite de frente."),
]


def hoja_competencia(wb):
    h = wb.create_sheet("Competencia")
    titulo_hoja(h, "Contra quién se compite, por mercado",
                "Los precios son órdenes de magnitud tomados de la comunicación pública de "
                "cada producto y cambian seguido: verificalos antes de usarlos en una "
                "propuesta comercial. La última columna es lo que hay que decir en una "
                "reunión cuando aparece ese nombre.")
    encabezado(h, 4, ["Mercado", "Competidor", "Qué hace", "Precio aproximado",
                      "Su fortaleza", "Su debilidad", "Cómo se le gana"],
               anchos=[11, 30, 46, 26, 40, 46, 56])
    fila = 5
    colores = {"Uruguay": "E8EDFA", "LATAM": "F0FDF4", "Mundo": "FFF7ED"}
    for mercado, nombre, que_hace, precio, fuerte, debil, gana in COMPETENCIA:
        h.cell(row=fila, column=1, value=mercado).font = NEGRITA
        h.cell(row=fila, column=2, value=nombre).font = NEGRITA
        for col, texto in ((3, que_hace), (4, precio), (5, fuerte), (6, debil), (7, gana)):
            h.cell(row=fila, column=col, value=texto).alignment = ARRIBA
        for col in range(1, 8):
            h.cell(row=fila, column=col).fill = PatternFill("solid", fgColor=colores[mercado])
            h.cell(row=fila, column=col).border = BORDE
        h.row_dimensions[fila].height = 62
        fila += 1

    fila += 1
    h.cell(row=fila, column=1, value="Dónde está el lugar de MV AutoML Studio").font = \
        Font(bold=True, color=AZUL, size=12)
    fila += 1
    for linea in [
        "Arriba está lo empresarial (DataRobot, H2O, Dataiku): potentísimo y a partir de "
        "cinco cifras al año. Abajo está lo gratuito (KNIME, Orange): capaz, pero hay que "
        "saber armarlo. En el medio quedan las herramientas web por suscripción, que son "
        "fáciles pero suben tus datos a la nube ajena.",
        "El hueco es ese: una PyME o un área de riesgo/cobranzas que quiere predecir, no "
        "tiene equipo de datos, no puede pagar cinco cifras y NO PUEDE subir los datos de "
        "sus clientes a un servidor de otro. Ahí no hay casi nadie a US$ 39 por mes.",
        "Los tres argumentos que hay que repetir siempre: (1) tus datos no salen de tu "
        "computadora; (2) te muestra el resultado sobre datos que el modelo nunca vio, así "
        "que el número no está inflado; (3) lo usa el que decide, no hace falta un "
        "científico de datos.",
        "El riesgo real no es DataRobot: es que la empresa siga usando la planilla de Excel "
        "de siempre. La venta se gana mostrando un caso propio de ellos, con sus datos, en "
        "la primera reunión.",
    ]:
        c = h.cell(row=fila, column=1, value="· " + linea)
        c.alignment = ARRIBA
        h.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=7)
        h.row_dimensions[fila].height = 44
        fila += 1
    return h


def main():
    wb = Workbook()
    wb.remove(wb.active)
    hoja_calificacion(wb)
    hoja_supuestos(wb)
    hoja_escenario(wb, "Sin publicidad", con_ads=False)
    hoja_escenario(wb, "Con publicidad", con_ads=True)
    hoja_resumen(wb)
    hoja_impuestos(wb)
    hoja_competencia(wb)
    wb.move_sheet("Resumen 1-24 meses", offset=-wb.index(wb["Resumen 1-24 meses"]))
    SALIDA.parent.mkdir(parents=True, exist_ok=True)
    wb.save(SALIDA)
    print(f"escrito: {SALIDA}")


if __name__ == "__main__":
    main()
