"""Almacén de datasets: ingesta sin límite de tamaño y consulta out-of-core.

Estrategia
----------
Cualquier origen (CSV, TXT, Excel, Parquet, consulta SQL) se normaliza a
Parquet particionado en disco. A partir de ahí DuckDB consulta el Parquet
*sin cargarlo en RAM*, de modo que el tamaño del dataset lo limita el disco
y no la memoria del proceso.

Formato en disco::

    data/datasets/<id>/
        meta.json          metadatos, esquema, origen
        part-0000.parquet  bloques de `settings.chunk_rows` filas
        part-0001.parquet
        ...
"""
from __future__ import annotations

import csv as _csv
import json
import os
import re
import shutil
import time
import uuid
from collections.abc import Iterable, Iterator
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

import duckdb
import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq

from ..config import settings

SAFE_ID = re.compile(r"^[A-Za-z0-9_\-]{1,64}$")
TEXT_EXT = {".csv", ".txt", ".tsv", ".dat", ".psv"}
EXCEL_EXT = {".xlsx", ".xlsm", ".xls", ".xlsb", ".ods"}
PARQUET_EXT = {".parquet", ".pq"}
JSON_EXT = {".json", ".jsonl", ".ndjson"}
SUPPORTED_EXT = TEXT_EXT | EXCEL_EXT | PARQUET_EXT | JSON_EXT


class IngestError(RuntimeError):
    """Error de ingesta legible por el usuario final."""


# ─────────────────────────────────────────────────────────────── metadatos ────
@dataclass
class DatasetMeta:
    id: str
    name: str
    source: str                      # upload | sql | derived
    origin: dict[str, Any] = field(default_factory=dict)
    rows: int = 0
    columns: list[dict[str, Any]] = field(default_factory=list)
    size_bytes: int = 0
    created_at: float = field(default_factory=time.time)
    parent_id: str | None = None
    notes: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        d = asdict(self)
        d["n_columns"] = len(self.columns)
        return d


# ───────────────────────────────────────────────────────── detección CSV ──────
def sniff_text(path: Path, sample_bytes: int = 256_000) -> dict[str, Any]:
    """Detecta encoding, separador y decimal de un archivo de texto.

    Se lee sólo una muestra del principio: funciona igual con 10 MB que con
    100 GB porque nunca recorre el archivo entero.

    El separador NO se elige por frecuencia. Un archivo con montos como
    ``$ 1.711,20`` tiene comas dentro de los valores, y elegir la coma partiría
    cada fila por la mitad corrompiendo los números en silencio. Se exige que el
    candidato produzca una cantidad de campos **estable** entre el encabezado y
    las filas de datos; si ninguno lo logra, el archivo es de una sola columna.
    """
    raw = path.open("rb").read(sample_bytes)
    encoding = "utf-8"
    try:
        raw.decode("utf-8-sig")
        encoding = "utf-8-sig" if raw[:3] == b"\xef\xbb\xbf" else "utf-8"
    except UnicodeDecodeError:
        try:
            import chardet  # opcional

            guess = chardet.detect(raw[:100_000]) or {}
            encoding = guess.get("encoding") or "latin-1"
        except Exception:
            encoding = "latin-1"

    text = raw.decode(encoding, errors="replace")
    lines = [ln for ln in text.splitlines() if ln.strip()]
    if len(lines) > 2:
        lines = lines[:-1]        # la última puede venir cortada por el muestreo
    lines = lines[:60]
    sep = _pick_separator(lines)

    # decimal: se decide por evidencia, no por el separador elegido
    body = "\n".join(lines[1:])
    if sep == ",":
        decimal = "."
    else:
        con_coma = len(re.findall(r"\d,\d", body))
        con_punto = len(re.findall(r"\d\.\d", body))
        decimal = "," if con_coma > con_punto else "."
    return {"encoding": encoding, "sep": sep, "decimal": decimal}


def _pick_separator(lines: list[str], candidates: str = ";,\t|") -> str:
    """Elige el separador que parte las filas de forma consistente."""
    if not lines:
        return ","
    mejor, mejor_puntaje = ",", -1.0
    for cand in candidates:
        try:
            filas = list(_csv.reader(lines, delimiter=cand))
        except Exception:
            continue
        if not filas:
            continue
        n_cabecera = len(filas[0])
        if n_cabecera < 2:
            continue
        datos = filas[1:] or filas
        iguales = sum(1 for f in datos if len(f) == n_cabecera)
        consistencia = iguales / len(datos)
        if consistencia < 0.9:
            continue
        # a igual consistencia gana el que produce más columnas: un separador
        # equivocado casi nunca parte parejo en muchas columnas
        puntaje = consistencia * 100 + n_cabecera
        if puntaje > mejor_puntaje:
            mejor, mejor_puntaje = cand, puntaje
    if mejor_puntaje < 0:
        # ningún candidato parte de forma estable: una sola columna. Se elige un
        # carácter que con seguridad no aparece, para no romper las filas.
        texto = "\n".join(lines)
        for cand in ("\x01", "|", "\t", ";"):
            if cand not in texto:
                return cand
        return ";"
    return mejor


def _clean_columns(cols: Iterable[Any]) -> list[str]:
    """Nombres de columna únicos, sin espacios sobrantes ni vacíos."""
    out, seen = [], {}
    for i, c in enumerate(cols):
        name = str(c).strip() if c is not None and str(c).strip() != "" else f"col_{i+1}"
        name = re.sub(r"\s+", " ", name)
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        out.append(name)
    return out


def _normalize_chunk(df: pd.DataFrame) -> pd.DataFrame:
    """Tipado consistente entre bloques.

    Sin esto, un bloque puede inferir int64 y el siguiente float64 sobre la
    misma columna, y el Parquet resultante queda con esquemas incompatibles.
    La regla: numéricos a float64, enteros a int64 sólo si no hay nulos,
    todo lo demás a string. Los tipos finos se deciden después, en el ETL.
    """
    for c in df.columns:
        s = df[c]
        if pd.api.types.is_bool_dtype(s):
            df[c] = s.astype("boolean").astype("object").where(s.notna(), None)
            df[c] = pd.to_numeric(df[c], errors="coerce")
        elif pd.api.types.is_integer_dtype(s):
            df[c] = s.astype("float64")
        elif pd.api.types.is_float_dtype(s):
            df[c] = s.astype("float64")
        elif pd.api.types.is_datetime64_any_dtype(s):
            df[c] = s
        else:
            df[c] = s.astype("string")
    return df


# ─────────────────────────────────────────────────────────── escritor ─────────
class _ParquetSink:
    """Escribe bloques a Parquet unificando el esquema entre bloques."""

    def __init__(self, folder: Path):
        self.folder = folder
        self.folder.mkdir(parents=True, exist_ok=True)
        self.n = 0
        self.rows = 0
        self.schema: pa.Schema | None = None

    def write(self, df: pd.DataFrame) -> None:
        if df is None or len(df) == 0:
            return
        df = _normalize_chunk(df.reset_index(drop=True))
        table = pa.Table.from_pandas(df, preserve_index=False)
        if self.schema is None:
            self.schema = table.schema
        elif table.schema != self.schema:
            table = _cast_to(table, self.schema)
        pq.write_table(table, self.folder / f"part-{self.n:04d}.parquet",
                       compression="zstd")
        self.n += 1
        self.rows += len(df)

    def close(self) -> pa.Schema | None:
        return self.schema


def _cast_to(table: pa.Table, schema: pa.Schema) -> pa.Table:
    """Alinea un bloque al esquema del primer bloque (columnas y tipos)."""
    cols = []
    for f in schema:
        if f.name in table.column_names:
            col = table.column(f.name)
            try:
                col = col.cast(f.type, safe=False)
            except (pa.ArrowInvalid, pa.ArrowNotImplementedError):
                col = col.cast(pa.string()) if f.type == pa.string() else pa.nulls(len(table), f.type)
        else:
            col = pa.nulls(len(table), f.type)
        cols.append(col)
    return pa.Table.from_arrays([pa.chunked_array(c) if not isinstance(c, pa.ChunkedArray) else c
                                 for c in cols], schema=schema)


# ─────────────────────────────────────────────────────────── ingesta ──────────
def _ingest_text(path: Path, sink: _ParquetSink, opts: dict[str, Any]) -> None:
    reader = pd.read_csv(
        path,
        sep=opts.get("sep", ","),
        encoding=opts.get("encoding", "utf-8"),
        decimal=opts.get("decimal", "."),
        thousands=opts.get("thousands") or None,
        chunksize=settings.chunk_rows,
        low_memory=False,
        on_bad_lines="skip",
        skipinitialspace=True,
        na_values=["", "NA", "N/A", "NULL", "null", "#N/A", "-", "nan", "NaN"],
        keep_default_na=True,
    )
    first = True
    for chunk in reader:
        if first:
            chunk.columns = _clean_columns(chunk.columns)
            cols = list(chunk.columns)
            first = False
        else:
            chunk.columns = cols
        sink.write(chunk)


def _ingest_excel(path: Path, sink: _ParquetSink, opts: dict[str, Any]) -> None:
    """Lee Excel en modo *read_only* fila a fila.

    openpyxl en modo read_only usa un parser SAX: la RAM no crece con el
    tamaño de la hoja, por lo que un .xlsx de millones de filas se ingesta
    igual que uno de mil.
    """
    ext = path.suffix.lower()
    sheet = opts.get("sheet")
    if ext in {".xls", ".xlsb", ".ods"}:
        # formatos que openpyxl no lee en streaming: pandas de una sola vez
        engine = {"xls": "xlrd", "xlsb": "pyxlsb", "ods": "odf"}[ext.lstrip(".")]
        try:
            df = pd.read_excel(path, sheet_name=sheet or 0, engine=engine)
        except Exception as exc:  # pragma: no cover - depende de extras
            raise IngestError(
                f"No se pudo leer {ext}: falta el motor '{engine}'. "
                f"Convertí el archivo a .xlsx o .csv. Detalle: {exc}"
            ) from exc
        df.columns = _clean_columns(df.columns)
        sink.write(df)
        return

    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = None
        buf: list[tuple] = []
        for row in rows:
            if header is None:
                if row is None or all(v is None for v in row):
                    continue
                header = _clean_columns(row)
                continue
            if row is None or all(v is None for v in row):
                continue
            buf.append(row[: len(header)] + (None,) * max(0, len(header) - len(row)))
            if len(buf) >= settings.chunk_rows:
                sink.write(pd.DataFrame(buf, columns=header))
                buf.clear()
        if buf:
            sink.write(pd.DataFrame(buf, columns=header))
        if header is None:
            raise IngestError("La hoja de Excel está vacía.")
    finally:
        wb.close()


def _ingest_parquet(path: Path, sink: _ParquetSink) -> None:
    pf = pq.ParquetFile(path)
    for batch in pf.iter_batches(batch_size=settings.chunk_rows):
        df = batch.to_pandas()
        df.columns = _clean_columns(df.columns)
        sink.write(df)


def _ingest_json(path: Path, sink: _ParquetSink) -> None:
    lines = path.suffix.lower() in {".jsonl", ".ndjson"}
    if not lines:
        head = path.open("rb").read(4096).lstrip()
        lines = head.startswith(b"{") and b"\n{" in path.open("rb").read(65536)
    if lines:
        reader = pd.read_json(path, lines=True, chunksize=settings.chunk_rows)
        first = True
        cols = None
        for chunk in reader:
            if first:
                chunk.columns = _clean_columns(chunk.columns)
                cols = list(chunk.columns)
                first = False
            else:
                chunk.columns = cols
            sink.write(chunk)
    else:
        df = pd.read_json(path)
        df.columns = _clean_columns(df.columns)
        sink.write(df)


# ────────────────────────────────────────────────────────────── API ───────────
def dataset_path(ds_id: str) -> Path:
    if not SAFE_ID.match(ds_id):
        raise IngestError(f"Identificador de dataset inválido: {ds_id!r}")
    return settings.dataset_dir / ds_id


def new_id(prefix: str = "ds") -> str:
    return f"{prefix}_{uuid.uuid4().hex[:12]}"


def _schema_to_columns(folder: Path) -> list[dict[str, Any]]:
    ds = pq.ParquetDataset(str(folder))
    return [{"name": f.name, "arrow_type": str(f.type)} for f in ds.schema]


def register_folder(ds_id: str, meta: DatasetMeta) -> DatasetMeta:
    folder = dataset_path(ds_id)
    files = sorted(folder.glob("part-*.parquet"))
    if not files:
        raise IngestError("La ingesta no produjo datos: el origen quedó vacío.")
    meta.rows = int(sum(pq.ParquetFile(f).metadata.num_rows for f in files))
    meta.columns = _schema_to_columns(folder)
    meta.size_bytes = int(sum(f.stat().st_size for f in files))
    (folder / "meta.json").write_text(
        json.dumps(meta.to_dict(), ensure_ascii=False, indent=2, default=str), encoding="utf-8"
    )
    return meta


def ingest_file(path: Path, name: str, source: str = "upload",
                origin: dict[str, Any] | None = None,
                opts: dict[str, Any] | None = None) -> DatasetMeta:
    """Convierte un archivo de cualquier tamaño en un dataset consultable."""
    ext = path.suffix.lower()
    if ext not in SUPPORTED_EXT:
        raise IngestError(
            f"Extensión no soportada: {ext or '(sin extensión)'}. "
            f"Aceptadas: {', '.join(sorted(SUPPORTED_EXT))}"
        )
    opts = dict(opts or {})
    ds_id = new_id()
    folder = dataset_path(ds_id)
    sink = _ParquetSink(folder)
    try:
        if ext in TEXT_EXT:
            detected = sniff_text(path)
            detected.update({k: v for k, v in opts.items() if v not in (None, "")})
            opts = detected
            _ingest_text(path, sink, opts)
        elif ext in EXCEL_EXT:
            _ingest_excel(path, sink, opts)
        elif ext in PARQUET_EXT:
            _ingest_parquet(path, sink)
        else:
            _ingest_json(path, sink)
    except IngestError:
        shutil.rmtree(folder, ignore_errors=True)
        raise
    except Exception as exc:
        shutil.rmtree(folder, ignore_errors=True)
        raise IngestError(f"No se pudo leer el archivo: {exc}") from exc

    meta = DatasetMeta(
        id=ds_id, name=name, source=source,
        origin={**(origin or {}), "file": path.name, "bytes": path.stat().st_size, **opts},
    )
    return register_folder(ds_id, meta)


def ingest_frames(frames: Iterator[pd.DataFrame], name: str, source: str,
                  origin: dict[str, Any] | None = None,
                  parent_id: str | None = None) -> DatasetMeta:
    """Ingesta desde un iterador de DataFrames (SQL, transformaciones ETL)."""
    ds_id = new_id()
    folder = dataset_path(ds_id)
    sink = _ParquetSink(folder)
    try:
        cols = None
        for df in frames:
            if df is None or len(df) == 0:
                continue
            if cols is None:
                df.columns = _clean_columns(df.columns)
                cols = list(df.columns)
            else:
                df.columns = cols[: len(df.columns)]
            sink.write(df)
    except Exception as exc:
        shutil.rmtree(folder, ignore_errors=True)
        raise IngestError(f"Falló la ingesta: {exc}") from exc
    meta = DatasetMeta(id=ds_id, name=name, source=source,
                       origin=origin or {}, parent_id=parent_id)
    return register_folder(ds_id, meta)


def load_meta(ds_id: str) -> DatasetMeta:
    f = dataset_path(ds_id) / "meta.json"
    if not f.exists():
        raise IngestError(f"Dataset inexistente: {ds_id}")
    d = json.loads(f.read_text(encoding="utf-8"))
    d.pop("n_columns", None)
    return DatasetMeta(**d)


def list_datasets() -> list[dict[str, Any]]:
    out = []
    for folder in sorted(settings.dataset_dir.glob("*/"), key=os.path.getmtime, reverse=True):
        f = folder / "meta.json"
        if f.exists():
            try:
                out.append(json.loads(f.read_text(encoding="utf-8")))
            except Exception:
                continue
    return out


def delete_dataset(ds_id: str) -> None:
    shutil.rmtree(dataset_path(ds_id), ignore_errors=True)


def glob_expr(ds_id: str) -> str:
    folder = dataset_path(ds_id).as_posix()
    return f"read_parquet('{folder}/part-*.parquet', union_by_name=true)"


def connect() -> duckdb.DuckDBPyConnection:
    con = duckdb.connect(database=":memory:")
    con.execute("PRAGMA threads=4")
    # sin límite de memoria estricto: DuckDB derrama a disco si hace falta
    con.execute(f"PRAGMA temp_directory='{(settings.data_dir / 'tmp').as_posix()}'")
    return con


def query(ds_id: str, sql: str, params: list | None = None) -> pd.DataFrame:
    """Ejecuta SQL sobre el dataset. `{t}` se reemplaza por la tabla física."""
    con = connect()
    try:
        return con.execute(sql.replace("{t}", glob_expr(ds_id)), params or []).df()
    finally:
        con.close()


def head(ds_id: str, n: int | None = None, offset: int = 0) -> pd.DataFrame:
    n = n or settings.preview_rows
    return query(ds_id, f"SELECT * FROM {{t}} LIMIT {int(n)} OFFSET {int(offset)}")


def load_frame(ds_id: str, columns: list[str] | None = None,
               max_rows: int | None = None, seed: int = 42) -> pd.DataFrame:
    """Carga en memoria, muestreando de forma reproducible si es demasiado grande."""
    meta = load_meta(ds_id)
    cols = "*" if not columns else ", ".join(f'"{c}"' for c in columns)
    limit = max_rows or settings.max_train_rows
    if meta.rows <= limit:
        return query(ds_id, f"SELECT {cols} FROM {{t}}")
    frac = limit / meta.rows
    con = connect()
    try:
        con.execute(f"SELECT setseed({(seed % 1000) / 1000.0})")
        return con.execute(
            f"SELECT {cols} FROM {glob_expr(ds_id)} USING SAMPLE {frac * 100:.6f}% (bernoulli)"
        ).df()
    finally:
        con.close()
