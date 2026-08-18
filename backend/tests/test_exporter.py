"""Exportación a Excel, CSV y Parquet."""
from __future__ import annotations

import pandas as pd
import pytest
from app.core import automl as A
from app.core import exporter, profiling
from openpyxl import load_workbook


@pytest.fixture(scope="module")
def informe(frame_binary):
    df = frame_binary.drop(columns=["id", "constante", "casi_vacia", "fecha", "monto_texto"])
    return A.train(df, A.TrainConfig(target="objetivo", budget_seconds=6, max_models=2,
                                     shap=False, permutation_importance=False))["report"]


def test_excel_trae_todas_las_hojas(tmp_root, informe, dataset_binary):
    prof = profiling.profile(dataset_binary.id)
    out = tmp_root / "informe.xlsx"
    info = exporter.build_report(out, report=informe, profile=prof,
                                 dataset_id=dataset_binary.id, data_limit=200)
    assert out.exists() and info["size_bytes"] > 5000
    wb = load_workbook(out, read_only=True)
    for hoja in ["Resumen Ejecutivo", "Comparativa de Modelos", "Analisis de Variables",
                 "Calidad de Datos", "Analisis Estadistico", "Datos"]:
        assert hoja in wb.sheetnames, wb.sheetnames
    wb.close()


def test_el_resumen_reporta_el_holdout(tmp_root, informe):
    out = tmp_root / "resumen.xlsx"
    exporter.build_report(out, report=informe)
    wb = load_workbook(out, read_only=True)
    ws = wb["Resumen Ejecutivo"]
    textos = [str(c.value) for fila in ws.iter_rows(max_row=6) for c in fila if c.value]
    assert any("holdout" in t.lower() for t in textos)
    wb.close()


def test_excel_sin_contenido_no_rompe(tmp_root):
    out = tmp_root / "vacio.xlsx"
    info = exporter.build_report(out)
    assert out.exists()
    assert info["sheets"] == []


def test_csv_conserva_todas_las_filas(dataset_binary):
    info = exporter.export_dataset(dataset_binary.id, "csv", sep=";", decimal=",")
    df = pd.read_csv(info["path"], sep=";", decimal=",", encoding="utf-8-sig")
    assert len(df) == dataset_binary.rows
    assert info["rows"] == dataset_binary.rows


def test_parquet_conserva_todas_las_filas(dataset_binary):
    info = exporter.export_dataset(dataset_binary.id, "parquet")
    df = pd.read_parquet(info["path"])
    assert len(df) == dataset_binary.rows


def test_limite_de_filas_se_respeta(dataset_binary):
    info = exporter.export_dataset(dataset_binary.id, "csv", limit=50)
    df = pd.read_csv(info["path"], sep=";", decimal=",", encoding="utf-8-sig")
    assert len(df) == 50


def test_valores_no_finitos_no_rompen_el_excel():
    assert exporter._v(float("nan")) == ""
    assert exporter._v(float("inf")) == ""
    assert exporter._v(None) == ""
    assert exporter._v(3.5) == 3.5
