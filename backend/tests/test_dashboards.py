"""Tablero automático y preguntas en lenguaje natural."""
from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from app.core import ask as ASK
from app.core import dashboards as D
from app.core import storage

EJEMPLOS = Path(__file__).resolve().parents[2] / "examples"


@pytest.fixture(scope="module")
def panel():
    return storage.ingest_file(EJEMPLOS / "cobranzas_panel.xlsx", "panel dash")


@pytest.fixture(scope="module")
def spec(panel):
    return D.detect_spec(panel.id)


def test_detecta_tiempo_metricas_y_dimensiones(spec):
    assert spec["time_column"] == "FechaObs"
    nombres = [m["name"] for m in spec["metrics"]]
    assert "TotalCobrado" in nombres
    assert spec["metrics"][0]["format"] == "money"      # el monto manda
    assert "Estado" in spec["dimensions"]
    assert any(k["kind"] == "rows" for k in spec["kpis"])
    assert any(c["type"] == "line" for c in spec["charts"])
    assert any(f["type"] == "daterange" for f in spec["filters"])


def test_kpi_filtrado_cuadra_contra_sql_manual(panel, spec):
    datos = D.run(panel.id, spec, {"Estado": ["Cobranza/Mora Temprana"]})
    kpi = next(k for k in datos["kpis"] if k.get("column") == "TotalCobrado")
    chk = storage.query(panel.id,
        "SELECT sum(TotalCobrado) v FROM {t} WHERE Estado = 'Cobranza/Mora Temprana'")
    assert abs(kpi["value"] - float(chk["v"].iloc[0])) < 1


def test_filtro_temporal(panel, spec):
    datos = D.run(panel.id, spec, {"FechaObs": {"from": "2025-01-01", "to": "2025-12-31"}})
    filas = next(k for k in datos["kpis"] if k["kind"] == "rows")
    chk = storage.query(panel.id, "SELECT count(*) n FROM {t} WHERE year(FechaObs) = 2025")
    assert int(filas["value"]) == int(chk["n"].iloc[0])


def test_delta_contra_periodo_anterior(panel, spec):
    datos = D.run(panel.id, spec, None)
    kpi = next(k for k in datos["kpis"] if k.get("column") == "TotalCobrado")
    assert "delta_pct" in kpi and kpi["delta_period"]


def test_filtro_desconocido_se_ignora_sin_romper(panel, spec):
    datos = D.run(panel.id, spec, {"ColumnaInventada": ["x"], "Estado": []})
    assert datos["table"]["total"] == panel.rows


def test_sin_inyeccion_por_valores_de_filtro(panel, spec):
    """Un valor malicioso viaja como parámetro, nunca concatenado."""
    datos = D.run(panel.id, spec, {"Estado": ["'; DROP TABLE x; --"]})
    assert datos["table"]["total"] == 0                 # no matchea nada, no rompe nada


def test_dataset_sin_fechas_ni_categorias(tmp_root):
    df = pd.DataFrame({"a": range(100), "b": [x * 1.5 for x in range(100)]})
    f = tmp_root / "plano.csv"
    df.to_csv(f, index=False)
    meta = storage.ingest_file(f, "plano")
    spec = D.detect_spec(meta.id)
    assert spec["time_column"] is None
    datos = D.run(meta.id, spec, None)
    assert datos["kpis"] and datos["table"]["total"] == 100


def test_export_xlsx_y_csv_cuadran(panel, spec):
    x = D.export(panel.id, spec, {"Estado": ["Comercial/Normal"]}, "xlsx")
    assert x["rows"] == 108     # 36 meses x 3 tipos
    import openpyxl
    wb = openpyxl.load_workbook(x["path"], read_only=True)
    assert "KPIs" in wb.sheetnames and "Datos" in wb.sheetnames
    wb.close()
    c = D.export(panel.id, spec, {"Estado": ["Comercial/Normal"]}, "csv")
    assert len(pd.read_csv(c["path"], sep=";")) == c["rows"] == 108


# ── preguntas en lenguaje natural (motor local, sin IA) ──────────────────────
@pytest.mark.parametrize("pregunta,fragmento_sql", [
    ("total de TotalCobrado por Estado", 'sum("TotalCobrado")'),
    ("cuántos registros hay", "count(*)"),
    ("promedio de MontoACobrarVencido", 'avg("MontoACobrarVencido")'),
    ("total cobrado por estado en 2025", "year"),
    ("top 3 por TotalCobrado", "LIMIT 3"),
])
def test_traductor_local(panel, pregunta, fragmento_sql):
    r = ASK.ask(panel.id, pregunta)
    assert fragmento_sql in r["sql"]
    assert r["answer"] and r["row_count"] >= 1
    assert r["engine"] == "reglas locales"


def test_respuesta_cuadra_contra_sql_manual(panel):
    r = ASK.ask(panel.id, "total de TotalCobrado por Estado")
    chk = storage.query(panel.id,
        "SELECT Estado, sum(TotalCobrado) v FROM {t} GROUP BY 1 ORDER BY v DESC")
    assert r["rows"][0]["grupo"] == chk["Estado"].iloc[0]
    assert abs(r["rows"][0]["valor"] - float(chk["v"].iloc[0])) < 1


def test_pregunta_ininteligible_da_guia(panel):
    with pytest.raises(ASK.AskError, match="Formas que entiendo"):
        ASK.ask(panel.id, "qué opinás del clima en Montevideo")


def test_el_guardian_corta_sql_de_escritura(panel):
    with pytest.raises(ASK.AskError):
        ASK._run_sql(panel.id, "DROP TABLE x")
    with pytest.raises(ASK.AskError):
        ASK._run_sql(panel.id, "SELECT 1; SELECT 2")


def test_endpoint_ask(client, panel):
    r = client.post("/api/ai/ask", json={"dataset_id": panel.id,
                                         "question": "cuántos registros hay"})
    assert r.status_code == 200
    assert "648" in r.json()["answer"].replace(".", "")
    r = client.post("/api/ai/ask", json={"dataset_id": panel.id, "question": "zzz qqq"})
    assert r.status_code == 422


def test_endpoint_dashboard(client, panel):
    sp = client.get(f"/api/dashboards/{panel.id}/spec").json()
    assert sp["time_column"] == "FechaObs"
    datos = client.post(f"/api/dashboards/{panel.id}/run",
                        json={"spec": sp, "filters": {"Estado": ["Comercial/Normal"]}}).json()
    assert datos["table"]["total"] == 108


# ── columnas que llegan como texto y en realidad no lo son ───────────────────

@pytest.fixture(scope="module")
def colocaciones():
    return storage.ingest_file(EJEMPLOS / "colocaciones_y_tasas.xlsx", "colocaciones dash")


def test_periodo_de_texto_sirve_como_eje_temporal(colocaciones):
    """«2023-01» guardado como texto es una serie de tiempo, no una categoría."""
    spec = D.detect_spec(colocaciones.id)
    assert spec["time_column"] == "Periodo"
    assert "Periodo" in spec["expressions"]
    datos = D.run(colocaciones.id, None, None)
    serie = next(c for c in datos["charts"] if c["type"] == "line")
    assert len(serie["data"]) == spec["rows"]           # un punto por período
    assert serie["data"][0]["x"] == "2023-01-01"


def test_monto_con_separador_de_miles_es_metrica(colocaciones):
    """«401.593.821» es el monto del negocio, no un texto sin sentido."""
    spec = D.detect_spec(colocaciones.id)
    assert "Colocacion" in [m["name"] for m in spec["metrics"]]
    datos = D.run(colocaciones.id, None, None)
    kpi = next(k for k in datos["kpis"] if k.get("column") == "Colocacion")
    assert kpi["value"] > 1e9                            # se sumó como número


def test_filtro_temporal_no_cuela_el_periodo_siguiente(colocaciones):
    """El día 'hasta' entra entero; la medianoche del siguiente no."""
    datos = D.run(colocaciones.id, None,
                  {"Periodo": {"from": "2025-01-01", "to": "2025-12-31"}})
    assert datos["kpis"][0]["value"] == 12               # los doce meses de 2025


def test_dias_se_promedian_en_vez_de_sumarse(tmp_root):
    df = pd.DataFrame({"DiasAtraso": [10, 20, 30, 40] * 25,
                       "Monto": [100.5, 200.25, 300.75, 400.0] * 25})
    f = tmp_root / "atrasos.csv"
    df.to_csv(f, index=False)
    meta = storage.ingest_file(f, "atrasos")
    spec = D.detect_spec(meta.id)
    dias = next(m for m in spec["metrics"] if m["name"] == "DiasAtraso")
    assert dias["agg"] == "avg"
    datos = D.run(meta.id, None, None)
    kpi = next(k for k in datos["kpis"] if k.get("column") == "DiasAtraso")
    assert kpi["value"] == pytest.approx(25.0)           # promedio, no 2.500


def test_el_spec_del_cliente_no_entra_en_el_sql(panel):
    """Aunque manden un spec armado a mano, el servidor usa el suyo."""
    veneno = {"time_column": None, "expressions": {}, "metrics": [], "dimensions": [],
              "kpis": [{"id": "x", "kind": "metric", "column": "TotalCobrado",
                        "agg": "sum", "format": "money"}],
              "charts": [{"id": "c", "type": "line", "x": "FechaObs", "y": "TotalCobrado",
                          "agg": "sum", "grain": "month') , (SELECT 1"}],
              "filters": [], "table": {"columns": ["Estado"], "page_size": 5}}
    datos = D.run(panel.id, veneno, None)
    assert datos["spec"]["time_column"] == "FechaObs"    # el del servidor, no el enviado
    assert len(datos["table"]["columns"]) > 1                # tampoco la tabla recortada
